import base64
import json
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from string import Template
from zoneinfo import ZoneInfo

import requests
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook


st.set_page_config(
    page_title="Ikos Dashboard",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
    <style>
    footer {visibility: hidden;}
    header {visibility: hidden;}
    #MainMenu {visibility: hidden;}
    .block-container {
        padding: 0 !important;
        max-width: 100% !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# -----------------------
# Config
# -----------------------
TIMEZONE = "Europe/Athens"
LOGO_PATH = "logo.png"
GREETING_FADE_SECONDS = 3

API_KEY = st.secrets.get("API_KEY", "")
QUOTES_API_KEY = st.secrets.get("QUOTES_API_KEY", "")

OFFICE_LOCATIONS = {
    "Thessaloniki": "Thessaloniki,GR",
}

PROPERTY_LOCATIONS = {
    "Halkidiki": "Polygyros,GR",
    "Corfu": "Kerkyra,GR",
    "Kos": "Kos, South Aegean, Greece",
    "Crete": "Heraklion,GR",
    "Marbella": "Marbella,ES",
    "Mallorca": "Palma,ES",
}

BANK_HOLIDAYS = [
    (date(2026, 1, 1), "New Year's Day"),
    (date(2026, 1, 6), "Θεοφάνεια"),
    (date(2026, 2, 23), "Καθαρά Δευτέρα"),
    (date(2026, 3, 25), "25η Μαρτίου"),
    (date(2026, 4, 13), "Δευτέρα του Πάσχα"),
    (date(2026, 5, 1), "Πρωτομαγιά"),
    (date(2026, 8, 15), "Κοίμηση της Θεοτόκου"),
    (date(2026, 10, 28), "28η Οκτωβρίου"),
    (date(2026, 12, 25), "Χριστούγεννα"),
    (date(2026, 12, 26), "2η μέρα Χριστουγέννων"),
]

GREEK_PROPERTIES = [
    {"name": "Ikos Oceania", "opening": date(2026, 3, 26), "closing": date(2026, 10, 31)},
    {"name": "Ikos Olivia", "opening": date(2026, 4, 2), "closing": date(2026, 10, 31)},
    {"name": "Ikos Dassia", "opening": date(2026, 4, 2), "closing": date(2026, 10, 31)},
    {"name": "Ikos Aria", "opening": date(2026, 4, 23), "closing": date(2026, 10, 31)},
    {"name": "Ikos Odisia", "opening": date(2026, 4, 23), "closing": date(2026, 10, 31)},
    {"name": "Ikos Kissamos", "opening": date(2026, 4, 30), "closing": date(2026, 10, 31)},
]

SPANISH_PROPERTIES = [
    {"name": "Ikos Andalusia", "opening": date(2026, 3, 26), "closing": date(2026, 11, 7)},
    {"name": "Ikos Porto Petro", "opening": date(2026, 4, 10), "closing": date(2026, 11, 7)},
]

DUETTO_LIVE_DATE = date(2026, 5, 5)
ADMIN_KEY = st.secrets.get("ADMIN_KEY", "")
OCCUPANCY_SNAPSHOT_PATH = "occupancy_snapshot.json"

OCCUPANCY_TAB_CONFIG = {
    "IOC": {
        "property_name": "Ikos Oceania",
        "range": "Q7:Q226",
        "avg_cell": "Q227",
    },
    "IOL": {
        "property_name": "Ikos Olivia",
        "range": "V6:V218",
        "avg_cell": "V220",
    },
    "IDA": {
        "property_name": "Ikos Dassia",
        "range": "X6:X218",
        "avg_cell": "X219",
    },
    "IOD": {
        "property_name": "Ikos Odisia",
        "range": "AA6:AA197",
        "avg_cell": "AA198",
    },
    "IAR": {
        "property_name": "Ikos Aria",
        "range": "S6:S197",
        "avg_cell": "S198",
    },
    "IKI": {
        "property_name": "Ikos Kissamos",
        "range": "AC6:AC190",
        "avg_cell": "AC191",
    },
    "IAN": {
        "property_name": "Ikos Andalusia",
        "range": "AD6:AD232",
        "avg_cell": "AD240",
    },
    "IPP": {
        "property_name": "Ikos Porto Petro",
        "range": "Y6:Y217",
        "avg_cell": "Y218",
    },
}

raw_birthdays = st.secrets.get("BIRTHDAYS", [])

BIRTHDAYS = [
    (b["name"], (b["month"], b["day"]))
    for b in raw_birthdays
]

# -----------------------
# Helpers
# -----------------------
def get_image_base64(path: str) -> str:
    file_path = Path(path)
    if not file_path.exists():
        return ""
    return base64.b64encode(file_path.read_bytes()).decode()


def get_logo_base64(path: str) -> str:
    return get_image_base64(path)


def get_weather_icon_svg(weather: str) -> str:
    weather = (weather or "").strip()

    icons = {
        "Clear": """
            <svg viewBox="0 0 24 24" aria-hidden="true">
              <circle cx="12" cy="12" r="4.2" fill="#F5B301"></circle>
              <g stroke="#F5B301" stroke-width="1.8" stroke-linecap="round">
                <line x1="12" y1="2.5" x2="12" y2="5.2"></line>
                <line x1="12" y1="18.8" x2="12" y2="21.5"></line>
                <line x1="2.5" y1="12" x2="5.2" y2="12"></line>
                <line x1="18.8" y1="12" x2="21.5" y2="12"></line>
                <line x1="5.2" y1="5.2" x2="7.1" y2="7.1"></line>
                <line x1="16.9" y1="16.9" x2="18.8" y2="18.8"></line>
                <line x1="16.9" y1="7.1" x2="18.8" y2="5.2"></line>
                <line x1="5.2" y1="18.8" x2="7.1" y2="16.9"></line>
              </g>
            </svg>
        """,
        "Clouds": """
            <svg viewBox="0 0 24 24" aria-hidden="true">
              <ellipse cx="10" cy="13.2" rx="5.2" ry="3.4" fill="#C8D0DF"></ellipse>
              <ellipse cx="14.8" cy="12.8" rx="4.5" ry="3.1" fill="#B5C0D3"></ellipse>
              <ellipse cx="7.2" cy="14.1" rx="3.2" ry="2.5" fill="#D6DDE9"></ellipse>
            </svg>
        """,
        "Rain": """
            <svg viewBox="0 0 24 24" aria-hidden="true">
              <ellipse cx="10" cy="10.8" rx="5.2" ry="3.4" fill="#C8D0DF"></ellipse>
              <ellipse cx="14.8" cy="10.4" rx="4.5" ry="3.1" fill="#B5C0D3"></ellipse>
              <g stroke="#4A90E2" stroke-width="1.8" stroke-linecap="round">
                <line x1="8" y1="15.2" x2="6.8" y2="18.2"></line>
                <line x1="12" y1="15.2" x2="10.8" y2="18.2"></line>
                <line x1="16" y1="15.2" x2="14.8" y2="18.2"></line>
              </g>
            </svg>
        """,
        "Drizzle": """
            <svg viewBox="0 0 24 24" aria-hidden="true">
              <ellipse cx="10" cy="10.8" rx="5.2" ry="3.4" fill="#C8D0DF"></ellipse>
              <ellipse cx="14.8" cy="10.4" rx="4.5" ry="3.1" fill="#B5C0D3"></ellipse>
              <g stroke="#67A7EF" stroke-width="1.5" stroke-linecap="round">
                <line x1="9" y1="15.5" x2="8.2" y2="17.4"></line>
                <line x1="13" y1="15.5" x2="12.2" y2="17.4"></line>
                <line x1="17" y1="15.5" x2="16.2" y2="17.4"></line>
              </g>
            </svg>
        """,
        "Thunderstorm": """
            <svg viewBox="0 0 24 24" aria-hidden="true">
              <ellipse cx="10" cy="10.8" rx="5.2" ry="3.4" fill="#C8D0DF"></ellipse>
              <ellipse cx="14.8" cy="10.4" rx="4.5" ry="3.1" fill="#B5C0D3"></ellipse>
              <polygon points="12,14.4 9.5,18.6 12.4,18.6 10.8,21.4 15.2,16.6 12.4,16.6 14,14.4" fill="#F5B301"></polygon>
            </svg>
        """,
        "Snow": """
            <svg viewBox="0 0 24 24" aria-hidden="true">
              <ellipse cx="10" cy="10.8" rx="5.2" ry="3.4" fill="#C8D0DF"></ellipse>
              <ellipse cx="14.8" cy="10.4" rx="4.5" ry="3.1" fill="#B5C0D3"></ellipse>
              <g stroke="#7FB7FF" stroke-width="1.4" stroke-linecap="round">
                <line x1="8" y1="15.4" x2="8" y2="18.2"></line>
                <line x1="6.6" y1="16.8" x2="9.4" y2="16.8"></line>
                <line x1="12.5" y1="15.4" x2="12.5" y2="18.2"></line>
                <line x1="11.1" y1="16.8" x2="13.9" y2="16.8"></line>
                <line x1="16.5" y1="15.4" x2="16.5" y2="18.2"></line>
                <line x1="15.1" y1="16.8" x2="17.9" y2="16.8"></line>
              </g>
            </svg>
        """,
        "Mist": """
            <svg viewBox="0 0 24 24" aria-hidden="true">
              <g stroke="#B8C2D1" stroke-width="1.8" stroke-linecap="round">
                <line x1="5" y1="8" x2="19" y2="8"></line>
                <line x1="3.5" y1="12" x2="17.5" y2="12"></line>
                <line x1="6.5" y1="16" x2="20.5" y2="16"></line>
              </g>
            </svg>
        """,
        "Fog": """
            <svg viewBox="0 0 24 24" aria-hidden="true">
              <g stroke="#B8C2D1" stroke-width="1.8" stroke-linecap="round">
                <line x1="5" y1="8" x2="19" y2="8"></line>
                <line x1="3.5" y1="12" x2="17.5" y2="12"></line>
                <line x1="6.5" y1="16" x2="20.5" y2="16"></line>
              </g>
            </svg>
        """,
        "Haze": """
            <svg viewBox="0 0 24 24" aria-hidden="true">
              <g stroke="#B8C2D1" stroke-width="1.8" stroke-linecap="round">
                <line x1="5" y1="8" x2="19" y2="8"></line>
                <line x1="3.5" y1="12" x2="17.5" y2="12"></line>
                <line x1="6.5" y1="16" x2="20.5" y2="16"></line>
              </g>
            </svg>
        """,
        "Unavailable": """
            <svg viewBox="0 0 24 24" aria-hidden="true">
              <circle cx="12" cy="12" r="4" fill="#D3D8E2"></circle>
            </svg>
        """,
    }

    return icons.get(
        weather,
        """
        <svg viewBox="0 0 24 24" aria-hidden="true">
          <circle cx="9" cy="9" r="3.6" fill="#F5B301"></circle>
          <ellipse cx="12" cy="13.2" rx="5.2" ry="3.4" fill="#C8D0DF"></ellipse>
          <ellipse cx="16.2" cy="12.9" rx="4.1" ry="2.8" fill="#B5C0D3"></ellipse>
        </svg>
        """,
    )


def get_weather_temp_class(temp_value) -> str:
    if temp_value is None:
        return "temp-unavailable"
    if temp_value <= 10:
        return "temp-cold"
    if temp_value <= 19:
        return "temp-mild"
    if temp_value <= 27:
        return "temp-warm"
    return "temp-hot"


def get_weather_condition_class(weather: str) -> str:
    weather = (weather or "").strip().lower()

    mapping = {
        "clear": "cond-clear",
        "clouds": "cond-clouds",
        "rain": "cond-rain",
        "drizzle": "cond-drizzle",
        "thunderstorm": "cond-thunderstorm",
        "snow": "cond-snow",
        "mist": "cond-mist",
        "fog": "cond-mist",
        "haze": "cond-mist",
        "unavailable": "cond-unavailable",
    }
    return mapping.get(weather, "cond-default")


@st.cache_data(ttl=600, show_spinner=False)
def fetch_weather(query: str, api_key: str) -> dict:
    url = "https://api.openweathermap.org/data/2.5/weather"
    params = {
        "q": query,
        "appid": api_key,
        "units": "metric",
    }
    response = requests.get(url, params=params, timeout=10)
    return {
        "status_code": response.status_code,
        "json": response.json(),
    }


def get_weather_for_city(query: str) -> dict:
    if not API_KEY:
        return {
            "temp": "—",
            "temp_value": None,
            "temp_class": get_weather_temp_class(None),
            "weather": "Unavailable",
            "condition_class": get_weather_condition_class("Unavailable"),
            "icon": get_weather_icon_svg("Unavailable"),
        }

    try:
        result = fetch_weather(query, API_KEY)
        status_code = result["status_code"]
        data = result["json"]

        if status_code != 200:
            return {
                "temp": "—",
                "temp_value": None,
                "temp_class": get_weather_temp_class(None),
                "weather": "Unavailable",
                "condition_class": get_weather_condition_class("Unavailable"),
                "icon": get_weather_icon_svg("Unavailable"),
            }

        temp = round(data["main"]["temp"])
        weather = data["weather"][0]["main"]
        icon = get_weather_icon_svg(weather)

        return {
            "temp": f"{temp}°C",
            "temp_value": temp,
            "temp_class": get_weather_temp_class(temp),
            "weather": weather,
            "condition_class": get_weather_condition_class(weather),
            "icon": icon,
        }

    except Exception:
        return {
            "temp": "—",
            "temp_value": None,
            "temp_class": get_weather_temp_class(None),
            "weather": "Unavailable",
            "condition_class": get_weather_condition_class("Unavailable"),
            "icon": get_weather_icon_svg("Unavailable"),
        }


def render_weather_rows(locations: dict, office: bool = False) -> str:
    rows = []
    for label, query in locations.items():
        info = get_weather_for_city(query)
        row_class = "office-row" if office else "weather-row"
        rows.append(
            f"""
            <div class="{row_class}">
                <div class="weather-left">
                    <div class="weather-city">{label}</div>
                    <div class="weather-condition {info["condition_class"]}">
                        <span class="weather-icon">{info["icon"]}</span>
                        <span>{info["weather"]}</span>
                    </div>
                </div>
                <div class="weather-temp {info["temp_class"]}">{info["temp"]}</div>
            </div>
            """
        )
    return "".join(rows)


def get_next_holiday(today_: date):
    future_holidays = [(d, name) for d, name in BANK_HOLIDAYS if d >= today_]
    if not future_holidays:
        return None, None, None
    next_date, next_name = min(future_holidays, key=lambda x: x[0])
    days_left = (next_date - today_).days
    return next_name, next_date, days_left


def get_weekend_indicator(today_: date):
    weekday = today_.weekday()

    if weekday >= 5:
        return {
            "title": "Weekend Indicator",
            "name": "Weekend",
            "days_text": "Today",
            "alert_class": "alert-weekend",
        }

    days_to_saturday = 5 - weekday

    if days_to_saturday == 1:
        text = "Tomorrow"
        alert_class = "alert-warning"
    else:
        text = f"{days_to_saturday} days"
        alert_class = "alert-normal"

    return {
        "title": "Weekend Indicator",
        "name": "Next weekend",
        "days_text": text,
        "alert_class": alert_class,
    }


def get_theme_colors(dark_mode: bool) -> dict:
    if dark_mode:
        return {
            "bg": "#081225",
            "text": "#EAF1FF",
            "muted": "#A9B8D0",
            "section_title": "#93A7C4",
            "divider": "#22324A",
            "weather_city": "#EAF1FF",
            "temp_mild": "#C7D2E3",
            "alert_normal": "#EAF1FF",
            "alert_warning": "#F59E0B",
            "alert_danger": "#FB7185",
            "alert_weekend": "#34D399",
            "logo_shadow": "0 2px 10px rgba(0,0,0,0.35)",
            "card_bg": "#101C31",
            "card_border": "#22324A",
            "card_subtle": "#93A7C4",
            "card_track": "#243247",
            "card_fill_1": "#3B82F6",
            "card_fill_2": "#60A5FA",
            "card_status_future": "#F59E0B",
            "card_status_live": "#34D399",
            "card_status_done": "#A9B8D0",
            "right_card_bg": "#101C31",
            "right_card_border": "#22324A",
            "right_card_shadow": "0 2px 10px rgba(0,0,0,0.18)",
        }

    return {
        "bg": "#FFFFFF",
        "text": "#2F3345",
        "muted": "#5F6675",
        "section_title": "#5F6B7A",
        "divider": "#E3E8F0",
        "weather_city": "#2F3345",
        "temp_mild": "#475569",
        "alert_normal": "#2F3345",
        "alert_warning": "#D97706",
        "alert_danger": "#C2410C",
        "alert_weekend": "#2E8B57",
        "logo_shadow": "none",
        "card_bg": "#FFFFFF",
        "card_border": "#E3E8F0",
        "card_subtle": "#6B7280",
        "card_track": "#E8EDF5",
        "card_fill_1": "#1F5FAE",
        "card_fill_2": "#4A90E2",
        "card_status_future": "#D97706",
        "card_status_live": "#2E8B57",
        "card_status_done": "#6B7280",
        "right_card_bg": "#FFFFFF",
        "right_card_border": "#E3E8F0",
        "right_card_shadow": "0 2px 10px rgba(15, 23, 42, 0.04)",
    }


def format_days_text(days_value: int) -> str:
    if days_value < 0:
        return "Live"
    if days_value == 1:
        return "1 day"
    return f"{days_value} days"


def format_short_date(d: date) -> str:
    return d.strftime("%d %b %Y")


def get_greeting(now: datetime) -> str:
    hour = now.hour
    weekday = now.weekday()  # Monday=0 ... Sunday=6

    weekday_messages = [
        ((0, 6), "Τι έγινε, έχουμε αϋπνίες?"),
        ((6, 8), "Νωρίς σήμερα..."),
        ((8, 12), "Καλημέρα!"),
        ((12, 16), "Καλησπέρα!"),
        ((16, 17), "Ετοίμαζε πράγματα σιγά σιγά..."),
        ((17, 20), "Ακόμα εδώ???"),
        ((20, 24), "Το έκαψες..."),
    ]

    saturday_messages = [
        ((0, 6), "Σάββατο ξημερώματα και είσαι εδώ;"),
        ((6, 8), "Σάββατο και τόσο νωρίς;"),
        ((8, 12), "Καλημέρα... για Σάββατο πάντα"),
        ((12, 16), "Σάββατο μεσημέρι, τι φάση;"),
        ((16, 17), "Άντε, μάζευε πράγματα σιγά σιγά..."),
        ((17, 20), "Σάββατο απόγευμα και ακόμα εδώ???"),
        ((20, 24), "Οκ, το παράκανες σήμερα..."),
    ]

    sunday_messages = [
        ((0, 6), "Κυριακή ξημερώματα... όλα καλά;"),
        ((6, 8), "Κυριακή και ξύπνησες από τώρα;"),
        ((8, 12), "Καλημέρα... όσο καλή μπορεί να είναι..."),
        ((12, 16), "Κυριακή μεσημέρι, αύριο πάλι απ’ την αρχή"),
        ((16, 17), "Σιγά σιγά τελειώνει το παραμύθι..."),
        ((17, 20), "Κυριακή απόγευμα και ακόμα εδώ???"),
        ((20, 24), "Αύριο δουλειά. Τα κεφάλια μέσα."),
    ]

    if weekday == 5:
        messages = saturday_messages
    elif weekday == 6:
        messages = sunday_messages
    else:
        messages = weekday_messages

    for (start_hour, end_hour), message in messages:
        if start_hour <= hour < end_hour:
            return message

    return "Καλημέρα!"


def get_property_progress(today_: date, opening: date, closing: date) -> tuple[int, str]:
    total_days = (closing - opening).days
    if total_days <= 0:
        return 0, "Unknown"

    if today_ < opening:
        return 0, "Not started"
    if today_ > closing:
        return 100, "Completed"

    elapsed_days = (today_ - opening).days
    progress = int((elapsed_days / total_days) * 100)
    progress = max(0, min(100, progress))
    return progress, "In season"


def render_property_cards(properties: list[dict], today_: date, occupancy_data: dict | None = None) -> str:
    cards = []
    occupancy_data = occupancy_data or {}

    for prop in properties:
        progress, status = get_property_progress(today_, prop["opening"], prop["closing"])

        if status == "Not started":
            status_class = "property-status-future"
        elif status == "Completed":
            status_class = "property-status-done"
        else:
            status_class = "property-status-live"

        occ = occupancy_data.get(prop["name"], {})
        occupancy_html = ""

        if occ:
            occupancy_html = f"""
                <div class="occupancy-box">
                    <div class="occupancy-title">Occupancy</div>
                    <div class="occupancy-grid">
                        <div class="occupancy-item">
                            <div class="occupancy-label">Min</div>
                            <div class="occupancy-value">{format_percent_display(occ.get("min"))}</div>
                        </div>
                        <div class="occupancy-item">
                            <div class="occupancy-label">Avg</div>
                            <div class="occupancy-value">{format_percent_display(occ.get("avg"))}</div>
                        </div>
                        <div class="occupancy-item">
                            <div class="occupancy-label">Max</div>
                            <div class="occupancy-value">{format_percent_display(occ.get("max"))}</div>
                        </div>
                    </div>
                </div>
            """

        cards.append(
            f"""
            <div class="property-card">
                <div class="property-card-name">{prop["name"]}</div>
                <div class="property-card-dates">
                    <span>{format_short_date(prop["opening"])}</span>
                    <span class="property-arrow">→</span>
                    <span>{format_short_date(prop["closing"])}</span>
                </div>
                <div class="property-card-status {status_class}">{status}</div>
                <div class="property-progress-bar">
                    <div class="property-progress-fill" style="width:{progress}%"></div>
                </div>
                <div class="property-progress-text">{progress}%</div>
                {occupancy_html}
                <div class="occupancy-value occ-min">{format_percent_display(occ.get("min"))}</div>
                <div class="occupancy-value occ-avg">{format_percent_display(occ.get("avg"))}</div>
                <div class="occupancy-value occ-max">{format_percent_display(occ.get("max"))}</div>
            </div>
            """
        )

    return "".join(cards)


def get_flag_svg(country: str) -> str:
    if country == "gr":
        return """
        <svg viewBox="0 0 28 20" aria-hidden="true">
          <rect width="28" height="20" rx="2" fill="#0D5EAF"></rect>
          <rect y="2.22" width="28" height="2.22" fill="#FFFFFF"></rect>
          <rect y="6.66" width="28" height="2.22" fill="#FFFFFF"></rect>
          <rect y="11.10" width="28" height="2.22" fill="#FFFFFF"></rect>
          <rect y="15.54" width="28" height="2.22" fill="#FFFFFF"></rect>
          <rect width="12" height="12" fill="#0D5EAF"></rect>
          <rect x="4.6" width="2.8" height="12" fill="#FFFFFF"></rect>
          <rect y="4.6" width="12" height="2.8" fill="#FFFFFF"></rect>
        </svg>
        """
    if country == "es":
        return """
        <svg viewBox="0 0 28 20" aria-hidden="true">
          <rect width="28" height="20" rx="2" fill="#AA151B"></rect>
          <rect y="5" width="28" height="10" fill="#F1BF00"></rect>
        </svg>
        """
    return ""
def get_days_until_next_birthday(today_: date) -> int:
    current_year = today_.year
    next_dates = []

    for _, (month, day) in BIRTHDAYS:
        bday_this_year = date(current_year, month, day)

        if bday_this_year >= today_:
            next_dates.append(bday_this_year)
        else:
            next_dates.append(date(current_year + 1, month, day))

    next_birthday = min(next_dates)
    return (next_birthday - today_).days


def get_today_birthdays(today_: date) -> list[str]:
    names = []
    for name, (month, day) in BIRTHDAYS:
        if today_.month == month and today_.day == day:
            names.append(name)
    return names

@st.cache_data(ttl=86400, show_spinner=False)
def fetch_quote_of_the_day(api_key: str) -> dict:
    fallback = {
        "quote": "Success is not final, failure is not fatal: it is the courage to continue that counts.",
        "author": "Winston Churchill",
    }

    if not api_key:
        return fallback

    url = "https://api.api-ninjas.com/v2/quoteoftheday"
    headers = {"X-Api-Key": api_key}

    try:
        response = requests.get(url, headers=headers, timeout=10)

        # Temporary debug
        print("QUOTE STATUS:", response.status_code)
        print("QUOTE TEXT:", response.text)

        if response.status_code != 200:
            return fallback

        data = response.json()

        # Handle either dict or list just in case
        if isinstance(data, list) and len(data) > 0:
            item = data[0]
        elif isinstance(data, dict):
            item = data
        else:
            return fallback

        quote = item.get("quote")
        author = item.get("author")

        if not quote or not author:
            return fallback

        return {
            "quote": quote,
            "author": author,
        }

    except Exception as e:
        print("QUOTE ERROR:", e)
        return fallback

def is_admin_mode() -> bool:
    if not ADMIN_KEY:
        return False
    return st.query_params.get("admin") == ADMIN_KEY


def normalize_percent_value(value):
    if value is None or value == "":
        return None

    if isinstance(value, str):
        raw = value.strip().replace("%", "").replace(",", ".")
        if not raw:
            return None
        try:
            num = float(raw)
        except ValueError:
            return None
        return num

    try:
        num = float(value)
    except (TypeError, ValueError):
        return None

    # Excel percentage cells often come as 0.4083 for 40.83%
    if 0 <= num <= 1:
        return num * 100

    return num


def extract_range_values(ws, cell_range: str) -> list[float]:
    values = []
    for row in ws[cell_range]:
        for cell in row:
            val = normalize_percent_value(cell.value)
            if val is not None:
                values.append(val)
    return values


def parse_sheet_date(sheet_name: str, prefix: str):
    pattern = rf"^{prefix}\s+(\d{{2}}\.\d{{2}}\.\d{{4}})$"
    match = re.match(pattern, sheet_name.strip())
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%d.%m.%Y").date()
    except ValueError:
        return None


def find_latest_sheet_for_prefix(workbook, prefix: str):
    candidates = []

    for sheet_name in workbook.sheetnames:
        parsed_date = parse_sheet_date(sheet_name, prefix)
        if parsed_date is not None:
            candidates.append((parsed_date, sheet_name))

    if not candidates:
        return None

    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def extract_occupancy_snapshot_from_excel(file_bytes: bytes) -> dict:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    snapshot = {}

    for prefix, cfg in OCCUPANCY_TAB_CONFIG.items():
        sheet_name = find_latest_sheet_for_prefix(wb, prefix)
        if not sheet_name:
            continue

        ws = wb[sheet_name]
        range_values = extract_range_values(ws, cfg["range"])
        avg_value = normalize_percent_value(ws[cfg["avg_cell"]].value)

        if not range_values and avg_value is None:
            continue

        snapshot[cfg["property_name"]] = {
            "min": round(min(range_values), 2) if range_values else None,
            "max": round(max(range_values), 2) if range_values else None,
            "avg": round(avg_value, 2) if avg_value is not None else None,
            "sheet": sheet_name,
        }

    return snapshot


def save_occupancy_snapshot(snapshot: dict) -> None:
    payload = {
        "updated_at": datetime.now(ZoneInfo(TIMEZONE)).isoformat(),
        "data": snapshot,
    }
    Path(OCCUPANCY_SNAPSHOT_PATH).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def load_occupancy_snapshot() -> dict:
    p = Path(OCCUPANCY_SNAPSHOT_PATH)
    if not p.exists():
        return {}

    try:
        payload = json.loads(p.read_text(encoding="utf-8"))
        return payload.get("data", {})
    except Exception:
        return {}


def format_percent_display(value) -> str:
    if value is None:
        return "—"
    return f"{value:.2f}%"

# -----------------------
# Toggle + intro state
# -----------------------
dark_mode = st.toggle("🌙 Dark mode", value=False)
theme = get_theme_colors(dark_mode)

is_admin = is_admin_mode()

is_admin = is_admin_mode()

if is_admin:
    with st.expander("Admin upload", expanded=True):
        uploaded_occupancy_file = st.file_uploader(
            "Upload latest occupancy Excel",
            type=["xlsx", "xlsm"],
            key="occupancy_uploader",
        )

        if uploaded_occupancy_file is not None:
            try:
                file_bytes = uploaded_occupancy_file.read()
                occupancy_snapshot = extract_occupancy_snapshot_from_excel(file_bytes)
                save_occupancy_snapshot(occupancy_snapshot)
                st.success("Occupancy snapshot updated.")
            except Exception as e:
                st.error(f"Could not process the file: {e}")

if "intro_shown" not in st.session_state:
    st.session_state.intro_shown = False

# -----------------------
# Time calculations
# -----------------------
now = datetime.now(ZoneInfo(TIMEZONE))
today = now.date()

greeting = get_greeting(now)
show_greeting = not st.session_state.intro_shown
st.session_state.intro_shown = True

greeting_overlay_html = ""
if show_greeting:
    greeting_overlay_html = f"""
    <div class="greeting-overlay">
        <div class="greeting-text">{greeting}</div>
    </div>
    """

greeting_delay = GREETING_FADE_SECONDS if show_greeting else 0

# -----------------------
# Left column cards
# -----------------------
holiday_name, holiday_date, holiday_days = get_next_holiday(today)

holiday_html = ""
if holiday_name is not None:
    holiday_html = f"""
    <div class="section info-section">
        <div class="section-title">Next Bank Holiday</div>
        <div class="info-name alert-weekend">{holiday_name}</div>
        <div class="info-days alert-weekend">{format_days_text(holiday_days)}</div>
    </div>
    """

weekend = get_weekend_indicator(today)

weekend_html = f"""
<div class="section info-section">
    <div class="section-title">{weekend["title"]}</div>
    <div class="info-name {weekend["alert_class"]}">{weekend["name"]}</div>
    <div class="info-days {weekend["alert_class"]}">{weekend["days_text"]}</div>
</div>
"""

# -----------------------
# Right column cards
# -----------------------
quote_info = fetch_quote_of_the_day(QUOTES_API_KEY)
quote_html = f"""
<div class="right-info-card quote-card">
    <div class="section-title">Quote of the day</div>
    <div class="quote-text">“{quote_info["quote"]}”</div>
    <div class="quote-author">— {quote_info["author"]}</div>
</div>
"""
# -----------------------
# Birthday cards
# -----------------------
days_to_next_bday = get_days_until_next_birthday(today)

birthday_countdown_html = f"""
<div class="right-info-card">
    <div class="section-title">Days until next birthday</div>
    <div class="info-days alert-normal">{format_days_text(days_to_next_bday)}</div>
</div>
"""

today_birthdays = get_today_birthdays(today)
birthday_mode_class = "birthday-mode" if today_birthdays else ""

birthday_today_html = ""
if today_birthdays:
    names = ", ".join(today_birthdays)

    birthday_today_html = f"""
    <div class="right-info-card birthday-card">
        <div class="section-title">🎉 Celebration</div>
        <div class="info-name alert-weekend">🎉 Happy Birthday {names}!</div>

        <div class="confetti"></div>
        <div class="confetti"></div>
        <div class="confetti"></div>
        <div class="confetti"></div>
        <div class="confetti"></div>
        <div class="confetti"></div>
    </div>
    """

# -----------------------
# Logo
# -----------------------
logo_html = ""
logo_b64 = get_logo_base64(LOGO_PATH)
if logo_b64:
    logo_html = f"""
        <div class="logo">
            <img src="data:image/png;base64,{logo_b64}" alt="Logo">
        </div>
    """

# -----------------------
# Weather HTML
# -----------------------
office_weather_html = render_weather_rows(OFFICE_LOCATIONS, office=True)
property_weather_html = render_weather_rows(PROPERTY_LOCATIONS, office=False)

# -----------------------
# Middle column HTML
# -----------------------
occupancy_data = load_occupancy_snapshot()

greek_properties_html = render_property_cards(GREEK_PROPERTIES, today, occupancy_data)
spanish_properties_html = render_property_cards(SPANISH_PROPERTIES, today, occupancy_data)

greek_flag_svg = get_flag_svg("gr")
spanish_flag_svg = get_flag_svg("es")

# -----------------------
# HTML
# -----------------------
birthday_section_divider = ""
if birthday_today_html:
    birthday_section_divider = '<div class="section-divider"></div>'
html_template = Template(
    """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    <style>
        html, body {
            margin: 0;
            padding: 0;
            height: 100%;
            overflow: hidden;
            background: $bg;
            font-family: 'Inter', Arial, Helvetica, sans-serif;
            color: $text;
        }

        body {
            position: relative;
        }

        .page {
            display: flex;
            width: 100%;
            height: 100vh;
            background: $bg;
            opacity: 0;
            transform: translateY(10px);
            animation: dashboardFadeInUp 0.9s ease forwards;
            animation-delay: ${greeting_delay}s;
        }

        .greeting-overlay {
            position: fixed;
            inset: 0;
            width: 100vw;
            height: 100vh;
            background: $bg;
            display: flex;
            align-items: center;
            justify-content: center;
            z-index: 9999;
            pointer-events: none;
            animation: greetingFadeOut 0.5s ease forwards;
            animation-delay: ${greeting_delay}s;
        }

        .greeting-text {
            font-size: 112px;
            font-weight: 800;
            line-height: 1;
            color: #111111;
            text-align: center;
            opacity: 0;
            animation: fadeInOut ${greeting_seconds}s ease-in-out forwards;
            transform-origin: center center;
            padding: 0 30px;
        }

        .left {
            width: 28%;
            min-width: 280px;
            padding: 24px 28px 20px 32px;
            box-sizing: border-box;
            display: flex;
            flex-direction: column;
            justify-content: flex-start;
            height: 100vh;
            overflow: hidden;
        }

        .middle {
            width: 44%;
            padding: 24px 24px 24px 24px;
            box-sizing: border-box;
            display: flex;
            flex-direction: column;
            justify-content: flex-start;
            height: 100vh;
            overflow-y: auto;
            overflow-x: hidden;
        }

        .right {
            width: 28%;
            min-width: 280px;
            padding: 24px 32px 20px 28px;
            box-sizing: border-box;
            display: flex;
            flex-direction: column;
            justify-content: flex-start;
            height: 100vh;
            overflow: hidden;
        }

        .middle-content {
            width: 100%;
            display: flex;
            flex-direction: column;
        }

        .logo {
            text-align: center;
            margin-bottom: 14px;
            flex: 0 0 auto;
        }

        .logo img {
            width: 210px;
            max-width: 60vw;
            height: auto;
            pointer-events: none;
            user-select: none;
            -webkit-user-drag: none;
            filter: drop-shadow($logo_shadow);
        }

        .section {
            margin-bottom: 16px;
        }

        .section-title {
            font-size: 13px;
            font-weight: 700;
            color: $section_title;
            text-transform: uppercase;
            letter-spacing: 0.7px;
            margin-bottom: 12px;
        }

        .section-divider {
            height: 1px;
            background: $divider;
            margin: 12px 0 14px 0;
        }

        .office-row,
        .weather-row {
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            gap: 16px;
            margin-bottom: 8px;
        }

        .weather-left {
            text-align: left;
        }

        .weather-city {
            font-size: 17px;
            font-weight: 600;
            line-height: 1.2;
            color: $weather_city;
        }

        .weather-condition {
            display: flex;
            align-items: center;
            gap: 6px;
            font-size: 13px;
            margin-top: 3px;
            font-weight: 500;
        }

        .weather-icon {
            display: inline-flex;
            width: 16px;
            height: 16px;
            flex: 0 0 16px;
        }

        .weather-icon svg {
            width: 16px;
            height: 16px;
            display: block;
        }

        .weather-temp {
            font-size: 20px;
            font-weight: 700;
            line-height: 1.1;
            white-space: nowrap;
        }

        .temp-cold {
            color: #2563EB;
            font-weight: 800;
        }

        .temp-mild {
            color: $temp_mild;
        }

        .temp-warm {
            color: #F59E0B;
        }

        .temp-hot {
            color: #DC2626;
            font-weight: 800;
        }

        .temp-unavailable {
            color: #9CA3AF;
        }

        .cond-clear {
            color: #F59E0B;
        }

        .cond-clouds {
            color: #7B8798;
        }

        .cond-rain {
            color: #2563EB;
        }

        .cond-drizzle {
            color: #60A5FA;
        }

        .cond-thunderstorm {
            color: #6D28D9;
        }

        .cond-snow {
            color: #93C5FD;
        }

        .cond-mist {
            color: #9CA3AF;
        }

        .cond-unavailable,
        .cond-default {
            color: #9CA3AF;
        }

        .info-section {
            margin-top: 4px;
        }

        .info-name {
            font-size: 18px;
            font-weight: 600;
            margin-bottom: 6px;
        }

        .info-days {
            font-size: 20px;
            font-weight: 700;
        }

        .alert-normal {
            color: $alert_normal;
        }

        .alert-warning {
            color: $alert_warning;
        }

        .alert-danger {
            color: $alert_danger;
        }

        .alert-weekend {
            color: $alert_weekend;
        }

        .properties-group {
            margin-bottom: 24px;
        }

        .properties-group-header {
            display: flex;
            align-items: center;
            gap: 10px;
            margin-bottom: 14px;
        }

        .properties-group-title {
            font-size: 14px;
            font-weight: 700;
            color: $section_title;
            text-transform: uppercase;
            letter-spacing: 0.7px;
        }

        .group-flag {
            width: 22px;
            height: 16px;
            display: inline-flex;
            flex: 0 0 auto;
        }

        .group-flag svg {
            width: 22px;
            height: 16px;
            display: block;
            border-radius: 2px;
            box-shadow: 0 0 0 1px rgba(0,0,0,0.06);
        }

        .properties-grid {
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 14px;
        }

        .property-card {
            background: $card_bg;
            border: 1px solid $card_border;
            border-radius: 18px;
            padding: 16px 16px 14px 16px;
            box-sizing: border-box;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.04);
        }

        .property-card-name {
            font-size: 18px;
            font-weight: 700;
            color: $text;
            margin-bottom: 8px;
            line-height: 1.2;
        }

        .property-card-dates {
            display: flex;
            align-items: center;
            gap: 7px;
            flex-wrap: wrap;
            font-size: 13px;
            font-weight: 600;
            color: $card_subtle;
            margin-bottom: 8px;
        }

        .property-arrow {
            opacity: 0.75;
        }

        .property-card-status {
            font-size: 13px;
            font-weight: 700;
            margin-bottom: 10px;
            text-transform: uppercase;
            letter-spacing: 0.4px;
        }

        .property-status-future {
            color: $card_status_future;
        }

        .property-status-live {
            color: $card_status_live;
        }

        .property-status-done {
            color: $card_status_done;
        }

        .property-progress-bar {
            width: 100%;
            height: 10px;
            background: $card_track;
            border-radius: 999px;
            overflow: hidden;
            margin-bottom: 8px;
        }

        .property-progress-fill {
            height: 100%;
            background: linear-gradient(90deg, $card_fill_1 0%, $card_fill_2 100%);
            border-radius: 999px;
            transition: width 0.6s ease;
        }

        .property-progress-text {
            font-size: 13px;
            font-weight: 700;
            color: $text;
            text-align: right;
        }

        .right-info-card {
            background: $right_card_bg;
            border: 1px solid $right_card_border;
            border-radius: 18px;
            padding: 16px;
            box-sizing: border-box;
            box-shadow: $right_card_shadow;
            margin-bottom: 2px;
        }

        @keyframes fadeInOut {
            0%   { opacity: 0; transform: scale(0.96); }
            20%  { opacity: 1; transform: scale(1); }
            80%  { opacity: 1; transform: scale(1); }
            100% { opacity: 0; transform: scale(1.02); }
        }

        @keyframes greetingFadeOut {
            to {
                opacity: 0;
                visibility: hidden;
            }
        }

        @keyframes dashboardFadeInUp {
            from {
                opacity: 0;
                transform: translateY(10px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }

        @media (max-width: 1400px) {
            .greeting-text {
                font-size: 96px;
            }

            .property-card-name {
                font-size: 16px;
            }
        }

        @media (max-width: 1200px) {
            .properties-grid {
                gap: 12px;
            }

            .property-card {
                padding: 14px 14px 12px 14px;
            }

            .property-card-name {
                font-size: 15px;
            }

            .property-card-dates {
                font-size: 12px;
            }
        }

        @media (max-width: 1000px) {
            html, body {
                overflow: auto;
            }

            .greeting-text {
                font-size: 68px;
                padding: 0 24px;
            }

            .page {
                flex-direction: column;
                height: auto;
                overflow: visible;
            }

            .left,
            .middle,
            .right {
                width: 100%;
                min-width: 100%;
                max-width: 100%;
                padding: 20px;
                height: auto;
                overflow: visible;
            }

            .properties-grid {
                grid-template-columns: 1fr;
            }
        }
        /* -----------------------
    Birthday Premium Effects
    ----------------------- */

    .birthday-card {
        position: relative;
        overflow: hidden;
        animation: birthdayPulse 2.5s ease-in-out infinite;
    }

    @keyframes birthdayPulse {
        0%   { box-shadow: 0 0 0 rgba(52, 211, 153, 0.0); }
        50%  { box-shadow: 0 0 18px rgba(52, 211, 153, 0.25); }
        100% { box-shadow: 0 0 0 rgba(52, 211, 153, 0.0); }
    }

    /* Confetti particles */
    .confetti {
        position: absolute;
        width: 6px;
        height: 6px;
        opacity: 0.7;
        border-radius: 2px;
        animation: confettiFall linear infinite;
    }

    .confetti:nth-child(1) { left: 10%; animation-duration: 3s; background: #F59E0B; }
    .confetti:nth-child(2) { left: 25%; animation-duration: 3.5s; background: #34D399; }
    .confetti:nth-child(3) { left: 40%; animation-duration: 2.8s; background: #60A5FA; }
    .confetti:nth-child(4) { left: 55%; animation-duration: 3.2s; background: #FB7185; }
    .confetti:nth-child(5) { left: 70%; animation-duration: 3.6s; background: #A78BFA; }
    .confetti:nth-child(6) { left: 85%; animation-duration: 2.9s; background: #FBBF24; }

    @keyframes confettiFall {
        0% {
            top: -10%;
            transform: translateY(0) rotate(0deg);
        }
        100% {
            top: 110%;
            transform: translateY(0) rotate(360deg);
        }
    }
    /* -----------------------
        Birthday Premium Effects
    ----------------------- */

    .birthday-card {
        position: relative;
        overflow: hidden;
        animation: birthdayPulse 2.5s ease-in-out infinite;
    }

    @keyframes birthdayPulse {
        0%   { box-shadow: 0 0 0 rgba(52, 211, 153, 0.0); }
        50%  { box-shadow: 0 0 18px rgba(52, 211, 153, 0.22); }
        100% { box-shadow: 0 0 0 rgba(52, 211, 153, 0.0); }
    }

    .confetti {
        position: absolute;
        width: 6px;
        height: 6px;
        opacity: 0.7;
        border-radius: 2px;
        animation: confettiFall linear infinite;
    }

    .confetti:nth-child(3) { left: 10%; animation-duration: 3s; background: #F59E0B; }
    .confetti:nth-child(4) { left: 25%; animation-duration: 3.5s; background: #34D399; }
    .confetti:nth-child(5) { left: 40%; animation-duration: 2.8s; background: #60A5FA; }
    .confetti:nth-child(6) { left: 55%; animation-duration: 3.2s; background: #FB7185; }
    .confetti:nth-child(7) { left: 70%; animation-duration: 3.6s; background: #A78BFA; }
    .confetti:nth-child(8) { left: 85%; animation-duration: 2.9s; background: #FBBF24; }

    @keyframes confettiFall {
        0% {
            top: -10%;
            transform: rotate(0deg);
        }
        100% {
            top: 110%;
            transform: rotate(360deg);
        }
    }

    /* whole dashboard subtle celebration mode */
    .page.birthday-mode {
        position: relative;
    }

    .page.birthday-mode::before {
        content: "";
        position: absolute;
        inset: 0;
        pointer-events: none;
        background:
            radial-gradient(circle at 15% 20%, rgba(251, 191, 36, 0.06), transparent 22%),
            radial-gradient(circle at 85% 25%, rgba(96, 165, 250, 0.05), transparent 24%),
            radial-gradient(circle at 70% 80%, rgba(52, 211, 153, 0.05), transparent 26%);
        animation: birthdayAmbient 6s ease-in-out infinite alternate;
    }

    @keyframes birthdayAmbient {
        from {
            opacity: 0.45;
        }
        to {
            opacity: 0.85;
        }
    }
    .quote-card {
        display: flex;
        flex-direction: column;
        gap: 8px;
    }

    .quote-text {
        font-size: 18px;
        line-height: 1.5;
        font-style: italic;
        font-weight: 500;
        color: $text;
    }

    .quote-author {
        font-size: 14px;
        color: $muted;
        font-weight: 600;
    }
    .occupancy-box {
        margin-top: 12px;
        padding-top: 10px;
        border-top: 1px solid $divider;
    }

    .occupancy-title {
        font-size: 11px;
        font-weight: 700;
        color: $section_title;
        text-transform: uppercase;
        letter-spacing: 0.6px;
        margin-bottom: 8px;
    }

    .occupancy-grid {
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 8px;
    }

    .occupancy-item {
        background: rgba(127, 127, 127, 0.06);
        border-radius: 10px;
        padding: 8px 6px;
        text-align: center;
    }

    .occupancy-label {
        font-size: 11px;
        font-weight: 600;
        color: $card_subtle;
        margin-bottom: 4px;
    }

    .occupancy-value {
        font-size: 14px;
        font-weight: 700;
        color: $text;
        line-height: 1.2;
    }
    </style>
</head>
<body>
    $greeting_overlay_html

    <div class="page $birthday_mode_class">
        <div class="left">
            <div class="section">
                <div class="section-title">Weather in our offices</div>
                $office_weather_html
            </div>

            <div class="section-divider"></div>

            <div class="section">
                <div class="section-title">Weather in our properties</div>
                $property_weather_html
            </div>

            <div class="section-divider"></div>

            $holiday_html

            <div class="section-divider"></div>

            $weekend_html
        </div>

        <div class="middle">
            <div class="middle-content">
                $logo_html

                <div class="properties-group">
                    <div class="properties-group-header">
                        <span class="properties-group-title">Greek Properties</span>
                        <span class="group-flag">$greek_flag_svg</span>
                    </div>
                    <div class="properties-grid">
                        $greek_properties_html
                    </div>
                </div>

                <div class="properties-group">
                    <div class="properties-group-header">
                        <span class="properties-group-title">Spanish Properties</span>
                        <span class="group-flag">$spanish_flag_svg</span>
                    </div>
                    <div class="properties-grid">
                        $spanish_properties_html
                    </div>
                </div>
            </div>
        </div>

        <div class="right">
            $quote_html

            <div class="section-divider"></div>

            $duetto_html

            $birthday_section_divider

            $ecommerce_html
        </div>
    </div>
</body>
</html>
"""
)

html = html_template.substitute(
    greeting_overlay_html=greeting_overlay_html,
    greeting_seconds=GREETING_FADE_SECONDS,
    greeting_delay=greeting_delay,
    office_weather_html=office_weather_html,
    property_weather_html=property_weather_html,
    holiday_html=holiday_html,
    weekend_html=weekend_html,
    duetto_html=birthday_countdown_html,
    ecommerce_html=birthday_today_html,
    logo_html=logo_html,
    greek_properties_html=greek_properties_html,
    spanish_properties_html=spanish_properties_html,
    greek_flag_svg=greek_flag_svg,
    spanish_flag_svg=spanish_flag_svg,
    bg=theme["bg"],
    text=theme["text"],
    muted=theme["muted"],
    section_title=theme["section_title"],
    divider=theme["divider"],
    weather_city=theme["weather_city"],
    temp_mild=theme["temp_mild"],
    alert_normal=theme["alert_normal"],
    alert_warning=theme["alert_warning"],
    alert_danger=theme["alert_danger"],
    alert_weekend=theme["alert_weekend"],
    logo_shadow=theme["logo_shadow"],
    card_bg=theme["card_bg"],
    card_border=theme["card_border"],
    card_subtle=theme["card_subtle"],
    card_track=theme["card_track"],
    card_fill_1=theme["card_fill_1"],
    card_fill_2=theme["card_fill_2"],
    card_status_future=theme["card_status_future"],
    card_status_live=theme["card_status_live"],
    card_status_done=theme["card_status_done"],
    right_card_bg=theme["right_card_bg"],
    right_card_border=theme["right_card_border"],
    right_card_shadow=theme["right_card_shadow"],
    birthday_mode_class=birthday_mode_class,
    quote_html=quote_html,
    birthday_section_divider=birthday_section_divider,
)

components.html(html, height=860, scrolling=False)